"""Create a simple XLSX template with python-docx/openpyxl"""
from openpyxl import Workbook
import json

wb = Workbook()
ws = wb.active
ws.title = "Cost Report"

# Headers
ws['A1'] = 'Cost Analysis Report'
ws['A2'] = 'Application:'
ws['B2'] = '{{appName}}'
ws['A3'] = 'Generated:'
ws['B3'] = '{{generatedAt}}'
ws['A4'] = 'Summary:'
ws['B4'] = '{{summary}}'

# Services table
ws['A6'] = 'Service'
ws['B6'] = 'Provider'
ws['C6'] = 'Category'
ws['D6'] = 'Monthly Cost'

wb.save('templates/xlsx-template.xlsx')
print('XLSX template saved')
